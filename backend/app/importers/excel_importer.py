# backend/app/importers/excel_importer.py
"""Import game asset data from Excel workbooks into PostgreSQL + Elasticsearch.

Supports two sheet types:
- Model sheets (module_type=1): character/NPC model assets
- Action sheets (module_type=3): animation/action module assets

Sheet names matching ``^[PMFA]\\d{0,3}`` are treated as model sheets.
The sheet named "动作模组" is treated as the action sheet.
Known non-data sheets (rules, statistics, etc.) are skipped.
"""
from __future__ import annotations

import json
import os
import re
import uuid
import zipfile
from collections.abc import Sequence

import asyncpg
import openpyxl

from app.importers.wps_image_extractor import extract_wps_images
from app.services.es_sync_service import build_es_doc, bulk_index

# ---------------------------------------------------------------------------
# Column mappings
# ---------------------------------------------------------------------------

COLUMN_MAP = {
    "资源完整路径": "resource_path",
    "截图": "_thumbnail",
    "物种": "species",
    "性别": "gender",
    "地域": "region",
    "势力": "faction",
    "职业": "profession",
    "体型": "body_type",
    "年龄": "age_group",
    "衣着": "clothing",
    "特征": "features",
    "专属NPC": "exclusive_npc",
    "备注": "remark",
}

ACTION_COLUMN_MAP = {
    "体型": "body_type",
    "动作模组": "action_module",
    "备注": "remark",
    "动作ID": "action_id",
    "动作类型": "action_type",
    "动作资源": "resource_path",
    "插槽": "slot_name",
    "插槽路径": "slot_path",
    "特效资源": "effect_path",
    "动作说明": "remark",
}

SKIP_SHEETS = {
    "通用规则",
    "进度统计",
    "需要新增的动作",
    "特效标签",
    "问题模型记录区",
    "WpsReserved_CellImgList",
}

MODEL_SHEETS = {
    "P080【完成】",
    "P081【完成】",
    "M2【完成】",
    "F2【完成】",
    "F1【完成】",
    "M1【完成】",
    "A",
    "P003【完成】",
    "P006【完成】",
    "P009【完成】",
    "P010【完成】",
    "P011【完成】",
    "P014【完成】",
    "P018【完成】",
    "P022【完成】",
    "P031【完成】",
    "P091【完成】",
    "P092【完成】",
}

ACTION_SHEETS = {"动作模组"}

MULTI_VALUE_FIELDS = {
    "region",
    "faction",
    "profession",
    "clothing",
    "features",
    "scene",
    "color",
    "description",
}

# ---------------------------------------------------------------------------
# Example / enum row detection
# ---------------------------------------------------------------------------

# The canonical example resource path used across all sheets' row 2 (or row 3
# in some sheets).  Any row whose resource_path normalizes to this value is an
# example row and must be skipped during data import.
_EXAMPLE_PATH_NORMALIZED = "data/source/Npc_source/P080/模型/P080001_HD.mdl"
_EXAMPLE_PATH_UPPER = _EXAMPLE_PATH_NORMALIZED.upper()

# Rows whose first cell starts with notes/instructions also need skipping.
_NOTES_PREFIXES = ("1.", "\n1.", "注意事项", "编辑器需求")


def _is_example_or_notes_row(resource_path_raw, row_values) -> bool:
    """Return True if this row is the example/enum row or a notes row.

    Detection heuristics:
    1. resource_path matches the canonical example path P080001_HD.mdl
    2. First cell starts with notes/instructions text
    3. Multiple tag cells contain newline-separated enum listings (≥3 cells
       each with ≥3 newlines — normal data rarely has that many)
    """
    if resource_path_raw:
        path = str(resource_path_raw).strip().replace("\\", "/").rstrip("\n").strip()
        if path.upper() == _EXAMPLE_PATH_UPPER:
            return True
        for prefix in _NOTES_PREFIXES:
            if path.startswith(prefix):
                return True

    # Count cells that look like enum listings (≥3 newline-separated values)
    enum_cell_count = 0
    for val in row_values:
        if val and isinstance(val, str) and val.count("\n") >= 3:
            enum_cell_count += 1
    if enum_cell_count >= 3:
        return True

    return False


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------


def classify_sheet(name: str) -> int | None:
    """Return module_type for a sheet name, or ``None`` to skip.

    Only whitelisted sheets are imported:
    - MODEL_SHEETS -> 1 (model)
    - ACTION_SHEETS -> 3 (action)
    - Everything else -> None (skip)
    """
    if name in MODEL_SHEETS:
        return 1
    if name in ACTION_SHEETS:
        return 3
    return None


def parse_multi_value(val) -> list[str]:
    """Split a cell value on newlines or `` / `` into a list of strings.

    Returns an empty list for ``None`` or empty strings.
    """
    if not val:
        return []
    s = str(val).strip()
    if not s:
        return []
    if "\n" in s:
        parts = s.split("\n")
    elif " / " in s:
        parts = s.split(" / ")
    else:
        parts = [s]
    return [p.strip() for p in parts if p.strip()]


def normalize_path(path: str) -> str:
    """Normalize a Windows-style resource path to forward slashes."""
    return path.strip().replace("\\", "/").rstrip("\n").strip()


# ---------------------------------------------------------------------------
# Main import routine
# ---------------------------------------------------------------------------


UPSERT_ASSETS_SQL = """
    INSERT INTO assets (module_type, name, resource_path, thumbnail_path, tags)
    SELECT * FROM UNNEST(
        $1::smallint[],
        $2::text[],
        $3::text[],
        $4::text[],
        $5::jsonb[]
    ) AS t(module_type, name, resource_path, thumbnail_path, tags)
    ON CONFLICT (module_type, resource_path)
    DO UPDATE SET thumbnail_path = COALESCE(EXCLUDED.thumbnail_path, assets.thumbnail_path),
                  tags = EXCLUDED.tags,
                  updated_at = NOW()
    RETURNING id, module_type, name, resource_path,
              thumbnail_path, tags, created_at, updated_at
"""


async def _flush_asset_batch(
    pool: asyncpg.Pool,
    rows: Sequence[tuple[int, str, str, str | None, str]],
    es_batch: list[dict],
) -> dict:
    """Upsert a batch of assets and append returned ES docs to *es_batch*."""
    if not rows:
        return {"success": 0}

    module_types = [row[0] for row in rows]
    names = [row[1] for row in rows]
    resource_paths = [row[2] for row in rows]
    thumbnail_paths = [row[3] for row in rows]
    tags_json = [row[4] for row in rows]

    async with pool.acquire() as conn:
        result_rows = await conn.fetch(
            UPSERT_ASSETS_SQL,
            module_types,
            names,
            resource_paths,
            thumbnail_paths,
            tags_json,
        )

    for row_result in result_rows:
        es_batch.append(build_es_doc(dict(row_result)))

    return {"success": len(result_rows)}


async def import_excel(
    filepath: str,
    pool: asyncpg.Pool,
    previews_dir: str,
    batch_size: int = 1000,
    progress_interval: int = 5000,
) -> dict:
    """Parse an Excel workbook and upsert rows into the *assets* table.

    Returns a summary dict with counts and any error details.
    """
    batch_id = str(uuid.uuid4())[:8]
    print(f"Excel import: opening workbook {filepath}", flush=True)
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    print(f"Excel import: workbook opened, sheets={len(wb.sheetnames)}", flush=True)
    print("Excel import: extracting WPS image mappings", flush=True)
    wps_images = extract_wps_images(filepath)
    print(
        f"Excel import: WPS image mappings ready, sheets={len(wps_images)}",
        flush=True,
    )

    os.makedirs(previews_dir, exist_ok=True)

    # Pre-open the xlsx as a zip for image extraction
    print("Excel import: opening workbook zip for thumbnails", flush=True)
    xlsx_zip = zipfile.ZipFile(filepath)

    stats = {"success": 0, "skipped": 0, "failed": 0, "es_sync_failed": 0, "images_extracted": 0}
    errors: list[dict] = []
    es_batch: list[dict] = []
    asset_batch: list[tuple[int, str, str, str | None, str]] = []
    processed = 0

    async def flush_assets() -> None:
        nonlocal asset_batch
        result = await _flush_asset_batch(pool, asset_batch, es_batch)
        stats["success"] += result["success"]
        asset_batch = []

    async def flush_es() -> None:
        nonlocal es_batch
        if not es_batch:
            return
        resp = await bulk_index(es_batch)
        if resp.get("errors"):
            for item in resp["items"]:
                if "error" in item.get("index", {}):
                    stats["es_sync_failed"] += 1
        es_batch = []

    def print_progress(sheet_name: str) -> None:
        print(
            "Excel progress: "
            f"sheet={sheet_name} processed={processed} "
            f"success={stats['success']} skipped={stats['skipped']} "
            f"failed={stats['failed']}",
            flush=True,
        )

    for sheet_name in wb.sheetnames:
        module_type = classify_sheet(sheet_name)
        if module_type is None:
            continue

        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_map = ACTION_COLUMN_MAP if module_type == 3 else COLUMN_MAP

        # Some model sheets have a resource path value in col 0 instead of
        # the header "资源完整路径".  Detect and fix.
        if (
            module_type == 1
            and headers
            and headers[0]
            and headers[0] not in col_map
            and ("/" in headers[0] or "\\" in headers[0])
        ):
            headers[0] = "资源完整路径"

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=3, values_only=True), start=3
        ):
            processed += 1
            try:
                # Skip example/enum rows and notes rows
                if _is_example_or_notes_row(row[0] if row else None, row):
                    stats["skipped"] += 1
                    continue

                mapped: dict = {}
                for col_idx, header in enumerate(headers):
                    if not header or header not in col_map:
                        continue
                    field = col_map[header]
                    val = row[col_idx] if col_idx < len(row) else None

                    if field == "_thumbnail":
                        continue

                    if field == "resource_path":
                        mapped["resource_path"] = (
                            normalize_path(str(val)) if val else ""
                        )
                        continue

                    if field == "action_id" and val is not None:
                        try:
                            mapped[field] = int(val)
                        except (ValueError, TypeError):
                            mapped[field] = None
                        continue

                    if field in MULTI_VALUE_FIELDS:
                        mapped[field] = parse_multi_value(val)
                    else:
                        mapped[field] = str(val).strip() if val else ""

                if not mapped.get("resource_path"):
                    stats["skipped"] += 1
                    continue

                resource_path = mapped.pop("resource_path")
                name = (
                    resource_path.rsplit("/", 1)[-1]
                    if "/" in resource_path
                    else resource_path
                )
                tags = {k: v for k, v in mapped.items() if v}
                tags["source_sheet"] = sheet_name

                # Extract thumbnail from WPS embedded images
                thumbnail_path = None
                sheet_images = wps_images.get(sheet_name, {})
                media_path = sheet_images.get(row_idx)
                if media_path:
                    ext = os.path.splitext(media_path)[1] or ".png"
                    thumb_filename = f"{os.path.splitext(name)[0]}{ext}"
                    thumb_full_path = os.path.join(previews_dir, thumb_filename)
                    if not os.path.exists(thumb_full_path):
                        try:
                            img_data = xlsx_zip.read(media_path)
                            with open(thumb_full_path, "wb") as f:
                                f.write(img_data)
                            stats["images_extracted"] += 1
                        except (KeyError, OSError):
                            thumb_filename = None
                    thumbnail_path = thumb_filename

                asset_batch.append(
                    (
                        module_type,
                        name,
                        resource_path,
                        thumbnail_path,
                        json.dumps(tags, ensure_ascii=False),
                    )
                )

                if len(asset_batch) >= batch_size:
                    await flush_assets()

                if len(es_batch) >= batch_size:
                    await flush_es()

                if processed % progress_interval == 0:
                    print_progress(sheet_name)

            except Exception as e:
                stats["failed"] += 1
                errors.append(
                    {"sheet": sheet_name, "row": row_idx, "error": str(e)}
                )

    await flush_assets()
    await flush_es()
    print_progress("done")

    wb.close()
    xlsx_zip.close()
    return {"batch_id": batch_id, **stats, "errors": errors[:50]}
