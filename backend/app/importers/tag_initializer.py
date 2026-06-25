# backend/app/importers/tag_initializer.py
import openpyxl
import asyncpg

COLUMN_MAP = {
    "物种": "species", "性别": "gender", "地域": "region",
    "势力": "faction", "职业": "profession", "体型": "body_type",
    "年龄": "age_group", "衣着": "clothing", "特征": "features",
    "专属NPC": "exclusive_npc",
}

SKIP_SHEETS = {"通用规则", "进度统计", "动作模组", "需要新增的动作",
               "特效标签", "问题模型记录区", "WpsReserved_CellImgList"}

# Values that are descriptions/instructions, not real tag values
_JUNK_VALUES = {"固定标签", "开放标签", "开放标签（后期合并后固定）"}


def _clean_values(raw: set[str]) -> set[str]:
    """Filter out junk values and composite newline-separated strings."""
    clean = set()
    for v in raw:
        v = v.strip()
        if not v or v in _JUNK_VALUES:
            continue
        if "\n" in v:
            continue
        clean.add(v)
    return clean


async def _sync_tag_values_from_db(
    pool: asyncpg.Pool,
    module_type: int,
    all_values: dict[str, set[str]] | None = None,
) -> dict[str, list[str]]:
    """Extract actual tag values from assets table and write into tag_values.

    Works for any module_type. Scans all enum_single/enum_multi definitions,
    queries distinct values from assets.tags JSONB, and upserts into tag_values.
    """
    if all_values is None:
        all_values = {}

    async with pool.acquire() as conn:
        defs = await conn.fetch(
            "SELECT id, field_name, field_type FROM tag_definitions WHERE module_type=$1",
            module_type,
        )
        for d in defs:
            fn = d["field_name"]
            ft = d["field_type"]
            if ft not in ("enum_single", "enum_multi"):
                continue
            if ft == "enum_multi":
                rows = await conn.fetch(
                    """SELECT DISTINCT
                         CASE jsonb_typeof(tags->$1)
                           WHEN 'array' THEN elem
                           ELSE tags->>$1
                         END AS val
                       FROM assets,
                       LATERAL jsonb_array_elements_text(
                         CASE jsonb_typeof(tags->$1)
                           WHEN 'array' THEN tags->$1
                           ELSE jsonb_build_array(tags->>$1)
                         END
                       ) AS elem
                       WHERE module_type=$2 AND tags ? $1""",
                    fn, module_type,
                )
            else:
                # enum_single: also handle the case where the value is stored as a JSON array
                # (legacy/mixed data) by unfolding arrays the same way as enum_multi.
                rows = await conn.fetch(
                    """SELECT DISTINCT
                         CASE jsonb_typeof(tags->$1)
                           WHEN 'array' THEN elem
                           ELSE tags->>$1
                         END AS val
                       FROM assets,
                       LATERAL jsonb_array_elements_text(
                         CASE jsonb_typeof(tags->$1)
                           WHEN 'array' THEN tags->$1
                           ELSE jsonb_build_array(tags->>$1)
                         END
                       ) AS elem
                       WHERE module_type=$2 AND tags ? $1
                         AND tags->>$1 IS NOT NULL AND tags->>$1 != ''""",
                    fn, module_type,
                )
            db_values = {r["val"] for r in rows if r["val"]}
            all_values.setdefault(fn, set()).update(db_values)

        # Clean all values
        for field_name in list(all_values.keys()):
            all_values[field_name] = _clean_values(all_values[field_name])

        # Clear old and re-insert
        for field_name, values in all_values.items():
            def_id = await conn.fetchval(
                "SELECT id FROM tag_definitions WHERE module_type=$1 AND field_name=$2",
                module_type, field_name,
            )
            if not def_id:
                continue
            await conn.execute(
                "DELETE FROM tag_values WHERE definition_id=$1", def_id,
            )
            for i, val in enumerate(sorted(values)):
                await conn.execute(
                    """INSERT INTO tag_values (definition_id, value, sort_order)
                       VALUES ($1, $2, $3)
                       ON CONFLICT (definition_id, value) DO NOTHING""",
                    def_id, val, i,
                )

    return {k: sorted(v) for k, v in all_values.items()}


async def extract_enum_values_from_excel(
    filepath: str, pool: asyncpg.Pool
) -> dict[str, list[str]]:
    """从 Excel 第 2 行（枚举行）提取各维度可选值，并结合数据库实际值，写入 tag_values"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    all_values: dict[str, set[str]] = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        enum_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]

        for col_idx, header in enumerate(headers):
            if not header or header not in COLUMN_MAP:
                continue
            field_name = COLUMN_MAP[header]
            cell_val = enum_row[col_idx] if col_idx < len(enum_row) else None
            if not cell_val:
                continue
            values = {v.strip() for v in str(cell_val).split("\n") if v.strip()}
            all_values.setdefault(field_name, set()).update(values)

    wb.close()

    return await _sync_tag_values_from_db(pool, module_type=1, all_values=all_values)


async def sync_model_tag_values(pool: asyncpg.Pool) -> dict[str, list[str]]:
    """从已导入的模型资产中提取标签值，写入 tag_values（module_type=1）"""
    return await _sync_tag_values_from_db(pool, module_type=1)


async def sync_effect_tag_values(pool: asyncpg.Pool) -> dict[str, list[str]]:
    """从已导入的特效资产中提取标签值，写入 tag_values（module_type=2）"""
    return await _sync_tag_values_from_db(pool, module_type=2)


async def sync_animator_tag_values(pool: asyncpg.Pool) -> dict[str, list[str]]:
    """从已导入的动作资产中提取标签值，写入 tag_values（module_type=3）"""
    return await _sync_tag_values_from_db(pool, module_type=3)


async def sync_icon_tag_values(pool: asyncpg.Pool) -> dict[str, list[str]]:
    """从已导入的图标资产中提取标签值，写入 tag_values（module_type=4）"""
    return await _sync_tag_values_from_db(pool, module_type=4)


async def sync_all_tag_values(pool: asyncpg.Pool) -> dict[int, dict[str, list[str]]]:
    """同步所有模块的标签值 — 用于启动时和批量操作"""
    result: dict[int, dict[str, list[str]]] = {}
    async with pool.acquire() as conn:
        modules = await conn.fetch(
            "SELECT DISTINCT module_type FROM tag_definitions ORDER BY module_type"
        )
        for row in modules:
            mod = row["module_type"]
            try:
                result[mod] = await _sync_tag_values_from_db(pool, module_type=mod)
            except Exception:
                result[mod] = {}
    return result
