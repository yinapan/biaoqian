"""Fast backfill of source_sheet tag from Excel without re-importing images.

Reads the Excel to map (module_type, resource_path) -> sheet_name, then
runs batch SQL updates. Much faster than a full re-import.
"""
import asyncio
import json
import os
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "backend"))

import asyncpg
import openpyxl


COLUMN_MAP_MODEL = {"资源完整路径": "resource_path"}
COLUMN_MAP_ACTION = {"动作资源": "resource_path"}

SKIP_SHEETS = {
    "通用规则", "进度统计", "需要新增的动作",
    "特效标签", "问题模型记录区", "WpsReserved_CellImgList",
}


def classify_sheet(name: str):
    if name in SKIP_SHEETS:
        return None
    if name == "动作模组":
        return 3
    if re.match(r"^[PMFA]\d{0,3}", name):
        return 1
    return None


def normalize_path(path: str) -> str:
    return path.strip().replace("\\", "/").rstrip("\n").strip()


async def main():
    excel_path = sys.argv[1] if len(sys.argv) > 1 else "资源标签对照表.xlsx"
    pg_url = "postgresql://biaoqiao:biaoqiao_dev@localhost:5432/biaoqiao"

    if not os.path.exists(excel_path):
        print(f"Excel not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    updates: list[tuple[int, str, str]] = []

    for sheet_name in wb.sheetnames:
        module_type = classify_sheet(sheet_name)
        if module_type is None:
            continue

        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_map = COLUMN_MAP_ACTION if module_type == 3 else COLUMN_MAP_MODEL

        path_col_idx = None
        for i, h in enumerate(headers):
            if h and h in col_map:
                path_col_idx = i
                break

        if path_col_idx is None:
            print(f"  Skip {sheet_name}: no path column found")
            continue

        count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            val = row[path_col_idx] if path_col_idx < len(row) else None
            if not val:
                continue
            rp = normalize_path(str(val))
            if not rp:
                continue
            updates.append((module_type, rp, sheet_name))
            count += 1

        print(f"  {sheet_name} (type={module_type}): {count} rows")

    wb.close()
    print(f"Total rows to update: {len(updates)}")

    pool = await asyncpg.create_pool(pg_url)
    updated = 0
    batch_size = 200

    for i in range(0, len(updates), batch_size):
        batch = updates[i:i + batch_size]
        async with pool.acquire() as conn:
            for module_type, resource_path, sheet_name in batch:
                result = await conn.execute(
                    """UPDATE assets
                       SET tags = jsonb_set(COALESCE(tags, '{}'::jsonb), '{source_sheet}', $3::jsonb),
                           updated_at = NOW()
                       WHERE module_type = $1 AND resource_path = $2
                         AND (tags->>'source_sheet' IS NULL OR tags->>'source_sheet' != $4)""",
                    module_type,
                    resource_path,
                    json.dumps(sheet_name, ensure_ascii=False),
                    sheet_name,
                )
                if result and result != "UPDATE 0":
                    updated += 1
        print(f"  Progress: {min(i + batch_size, len(updates))}/{len(updates)} checked, {updated} updated")

    await pool.close()
    print(f"Done. {updated} records updated with source_sheet.")


if __name__ == "__main__":
    asyncio.run(main())
