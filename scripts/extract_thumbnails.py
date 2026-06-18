"""Extract WPS embedded images from Excel and update thumbnail_path in the database.

Standalone script — does NOT reload openpyxl (which is slow on 1GB+ files).
Uses the WPS image extractor + openpyxl only for header/row mapping.

Usage: python scripts/extract_thumbnails.py [--xlsx PATH] [--previews-dir DIR] [--pg-url URL]
"""
import argparse
import asyncio
import os
import re
import sys
import time
import zipfile

sys.path.insert(0, str(os.path.join(os.path.dirname(__file__), "..", "backend")))

import asyncpg
import openpyxl

from app.importers.wps_image_extractor import extract_wps_images


COLUMN_MAP = {
    "资源完整路径": "resource_path",
}

ACTION_COLUMN_MAP = {
    "动作资源": "resource_path",
}

SKIP_SHEETS = {
    "通用规则",
    "进度统计",
    "需要新增的动作",
    "特效标签",
    "问题模型记录区",
    "WpsReserved_CellImgList",
}


def classify_sheet(name: str) -> int | None:
    if name in SKIP_SHEETS:
        return None
    if name == "动作模组":
        return 3
    if re.match(r"^[PMFA]\d{0,3}", name):
        return 1
    return None


async def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default="资源标签对照表.xlsx")
    parser.add_argument("--previews-dir", default="previews")
    parser.add_argument("--pg-url", default="postgresql://biaoqiao:biaoqiao_dev@localhost:5432/biaoqiao")
    args = parser.parse_args()

    print(f"[1/4] Extracting WPS image mappings...")
    t0 = time.time()
    wps_images = extract_wps_images(args.xlsx)
    total_mappings = sum(len(v) for v in wps_images.values())
    print(f"       Found {total_mappings} image mappings in {time.time()-t0:.1f}s")

    print(f"[2/4] Extracting images from xlsx to {args.previews_dir}/...")
    os.makedirs(args.previews_dir, exist_ok=True)

    t0 = time.time()
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)

    # Build row_idx -> (name, media_path) mapping per sheet
    updates: list[tuple[int, str, str]] = []  # (module_type, resource_path_name, thumb_filename)
    extracted = 0
    skipped_exists = 0

    with zipfile.ZipFile(args.xlsx) as zf:
        for sheet_name in wb.sheetnames:
            module_type = classify_sheet(sheet_name)
            if module_type is None:
                continue

            sheet_images = wps_images.get(sheet_name, {})
            if not sheet_images:
                continue

            ws = wb[sheet_name]
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            col_map = ACTION_COLUMN_MAP if module_type == 3 else COLUMN_MAP

            # Find resource_path column index
            rp_col_idx = None
            for idx, header in enumerate(headers):
                if header and header in col_map and col_map[header] == "resource_path":
                    rp_col_idx = idx
                    break

            if rp_col_idx is None:
                continue

            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                media_path = sheet_images.get(row_idx)
                if not media_path:
                    continue

                rp_val = row[rp_col_idx] if rp_col_idx < len(row) else None
                if not rp_val:
                    continue

                resource_path = str(rp_val).strip().replace("\\", "/")
                name = resource_path.rsplit("/", 1)[-1] if "/" in resource_path else resource_path

                ext = os.path.splitext(media_path)[1] or ".png"
                thumb_filename = f"{os.path.splitext(name)[0]}{ext}"
                thumb_full_path = os.path.join(args.previews_dir, thumb_filename)

                if not os.path.exists(thumb_full_path):
                    try:
                        img_data = zf.read(media_path)
                        with open(thumb_full_path, "wb") as f:
                            f.write(img_data)
                        extracted += 1
                    except (KeyError, OSError) as e:
                        print(f"       [WARN] Failed to extract {media_path}: {e}")
                        continue
                else:
                    skipped_exists += 1

                updates.append((module_type, name, thumb_filename))

    wb.close()
    print(f"       Extracted {extracted} images, {skipped_exists} already existed ({time.time()-t0:.1f}s)")

    print(f"[3/4] Updating {len(updates)} thumbnail_path entries in database...")
    t0 = time.time()
    pool = await asyncpg.create_pool(args.pg_url)
    updated = 0
    try:
        async with pool.acquire() as conn:
            for module_type, name, thumb_filename in updates:
                result = await conn.execute(
                    """UPDATE assets SET thumbnail_path = $1, updated_at = NOW()
                       WHERE module_type = $2 AND name = $3 AND thumbnail_path IS NULL""",
                    thumb_filename, module_type, name,
                )
                if result.endswith("1"):
                    updated += 1
    finally:
        await pool.close()
    print(f"       Updated {updated} rows ({time.time()-t0:.1f}s)")

    print(f"[4/4] Triggering ES reindex...")
    admin_key = None
    env_path = os.path.join(os.path.dirname(__file__), "..", ".env")
    if os.path.exists(env_path):
        for line in open(env_path).read().splitlines():
            if line.startswith("ADMIN_API_KEY="):
                admin_key = line.split("=", 1)[1].strip()
                break
    if not admin_key:
        admin_key = "dev-admin-key-change-in-prod"
    if not admin_key or admin_key == "dev-admin-key-change-in-prod":
        print("WARNING: ADMIN_API_KEY is using default value! "
              "Set a secure key in .env for shared deployments.", file=sys.stderr)

    import urllib.request
    import json
    try:
        req = urllib.request.Request(
            "http://localhost/api/v1/admin/reindex-es",
            method="POST",
            headers={"X-Admin-Key": admin_key},
        )
        with urllib.request.urlopen(req, timeout=300) as resp:
            result = json.loads(resp.read())
        print(f"       Reindex done: {result}")
    except Exception as e:
        print(f"       [WARN] Reindex failed: {e}")
        print("       Run manually: curl -X POST http://localhost/api/v1/admin/reindex-es -H 'X-Admin-Key: <key>'")

    print("Done!")


if __name__ == "__main__":
    asyncio.run(main())
