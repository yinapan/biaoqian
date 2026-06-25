"""Import only the F1 sheet from the Excel workbook.

The F1 sheet has a quirky col-0 header (a path value, not "资源完整路径").
This script handles that and imports the ~670 F1 records with thumbnails.
"""
import asyncio
import json
import os
import re
import sys
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "backend"))

import asyncpg
import openpyxl

from app.importers.wps_image_extractor import extract_wps_images

COLUMN_MAP = {
    "资源完整路径": "resource_path",
    "截图": "_thumbnail",
    "物种": "species",
    "性别": "gender",
    "地域": "region",
    "势力": "faction",
    "职业": "profession",
    "体型": "body_type",
    "年龄": "age_group",
    "衣着": "clothing",
    "特征": "features",
    "专属NPC": "exclusive_npc",
    "备注": "remark",
}

MULTI_VALUE_FIELDS = {
    "region", "faction", "profession", "clothing", "features",
    "scene", "color", "description",
}

SHEET_NAME = "F1【完成】"
MODULE_TYPE = 1


def parse_multi_value(val) -> list[str]:
    if not val:
        return []
    s = str(val).strip()
    if not s:
        return []
    if "\n" in s:
        parts = s.split("\n")
    elif " / " in s:
        parts = s.split(" / ")
    else:
        parts = [s]
    return [p.strip() for p in parts if p.strip()]


def normalize_path(path: str) -> str:
    return path.strip().replace("\\", "/").rstrip("\n").strip()


async def main():
    excel_path = sys.argv[1] if len(sys.argv) > 1 else "资源标签对照表.xlsx"
    pg_url = "postgresql://biaoqiao:biaoqiao_dev@localhost:5432/biaoqiao"
    project_root = Path(__file__).resolve().parent.parent
    previews_dir = str(project_root / "runtime_data" / "model" / "previews")
    os.makedirs(previews_dir, exist_ok=True)

    print(f"Extracting WPS images from {excel_path}...")
    wps_images = extract_wps_images(excel_path)
    sheet_images = wps_images.get(SHEET_NAME, {})
    print(f"  Found {len(sheet_images)} images for {SHEET_NAME}")

    xlsx_zip = zipfile.ZipFile(excel_path)

    print(f"Reading {SHEET_NAME}...")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Fix quirky header: col 0 contains a path value instead of "资源完整路径"
    if headers[0] and ("/" in headers[0] or "\\" in headers[0]):
        headers[0] = "资源完整路径"

    pool = await asyncpg.create_pool(pg_url)
    stats = {"success": 0, "skipped": 0, "failed": 0, "images": 0}

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        try:
            mapped = {}
            for col_idx, header in enumerate(headers):
                if not header or header not in COLUMN_MAP:
                    continue
                field = COLUMN_MAP[header]
                val = row[col_idx] if col_idx < len(row) else None
                if field == "_thumbnail":
                    continue
                if field == "resource_path":
                    mapped["resource_path"] = normalize_path(str(val)) if val else ""
                    continue
                if field in MULTI_VALUE_FIELDS:
                    mapped[field] = parse_multi_value(val)
                else:
                    mapped[field] = str(val).strip() if val else ""

            if not mapped.get("resource_path"):
                stats["skipped"] += 1
                continue

            resource_path = mapped.pop("resource_path")
            name = resource_path.rsplit("/", 1)[-1] if "/" in resource_path else resource_path
            tags = {k: v for k, v in mapped.items() if v}
            tags["source_sheet"] = SHEET_NAME

            # Extract thumbnail
            thumbnail_path = None
            media_path = sheet_images.get(row_idx)
            if media_path:
                ext = os.path.splitext(media_path)[1] or ".png"
                thumb_filename = f"{os.path.splitext(name)[0]}{ext}"
                thumb_full_path = os.path.join(previews_dir, thumb_filename)
                if not os.path.exists(thumb_full_path):
                    try:
                        img_data = xlsx_zip.read(media_path)
                        with open(thumb_full_path, "wb") as f:
                            f.write(img_data)
                        stats["images"] += 1
                    except (KeyError, OSError):
                        thumb_filename = None
                thumbnail_path = f"model/{thumb_filename}" if thumb_filename else None

            async with pool.acquire() as conn:
                await conn.execute(
                    """INSERT INTO assets (module_type, name, resource_path, thumbnail_path, tags)
                       VALUES ($1, $2, $3, $4, $5::jsonb)
                       ON CONFLICT (module_type, resource_path)
                       DO UPDATE SET thumbnail_path = COALESCE($4, assets.thumbnail_path),
                                     tags = $5::jsonb, updated_at = NOW()""",
                    MODULE_TYPE, name, resource_path, thumbnail_path,
                    json.dumps(tags, ensure_ascii=False),
                )
                stats["success"] += 1

        except Exception as e:
            stats["failed"] += 1
            print(f"  Error row {row_idx}: {e}")

    await pool.close()
    wb.close()
    xlsx_zip.close()
    print(f"Done: {stats}")


if __name__ == "__main__":
    asyncio.run(main())
